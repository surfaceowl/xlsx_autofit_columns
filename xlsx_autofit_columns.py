"""
Python helper to update the format of an existing excel file output with
column widths for each column set to accommodate the longest content in any row, for all columns.

Ignores null cells to avoid zero-width columns in rare cases where intermediate columns are empty
Ignores formulas for setting length - since long formulas create very wide columns, and results from
formulas are often much narrower than the actual formula
"""

import openpyxl  # r/w excel 2010 and newer files
from openpyxl.utils import get_column_letter  # lookup column title from index num
import sys  # for taking file parameter as input


class XLSXAutoFitColumns(object):
    """
    CLASS to ingest an excel file and update the existing file with
    helps the user more easily read content when
    column widths are narrower than column content
    """

    def __init__(self, inputfile):
        self.inputfile = inputfile
        self.output_filename = inputfile
        self.workbook = openpyxl.load_workbook(inputfile)

    def fit_column_widths_for_one_sheet(self, worksheet):
        """
        Method to set the width parameter of each column in the excel file
        :param worksheet: worksheet we are processing
        :return:  worksheet object, with column widths set in excel
        """
        # cell_widths: dict of column numbers and MAX len of longest cell contents for each column
        cell_widths = {}
        for column in worksheet:
            for cell in column:
                if cell.value:  # skip empty cells
                    if (
                        str(cell.value)[0] == "="
                    ):  # skip formulas, often much longer than content
                        continue
                    cell_widths[cell.column] = max(
                        (cell_widths.get(cell.column, 0), len(str(cell.value)))
                    )

        for col, column_width in cell_widths.items():
            column_width = str(column_width)
            worksheet.column_dimensions[get_column_letter(col)].width = column_width

        return worksheet

    def process_all_worksheets(self) -> object:
        for worksheet in self.workbook.sheetnames:
            self.fit_column_widths_for_one_sheet(self.workbook[worksheet])
        self.save()
        return True

    def save(self):
        """
        Save the workbook
        :return: saved excel file, returns `True`
        """
        self.workbook.save(self.output_filename)
        return True


if __name__ == "__main__":

    # check if user has input file to fix, or use the default
    if len(sys.argv) == 1:
        inputfile = "sample_excel_data.xlsx"
    else:
        inputfile = sys.argv[1]

    fix_worksheet = XLSXAutoFitColumns(inputfile)
    fix_worksheet.process_all_worksheets()
